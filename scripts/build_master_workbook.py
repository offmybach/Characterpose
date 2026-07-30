#!/usr/bin/env python3
"""Build CGB_MASTER_outreach.xlsx from every letters-*.md source."""
import re, unicodedata, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

os.chdir('/home/user/Characterpose')

def n(s):
    if s is None: return ""
    return re.sub(r'[ \t]+', ' ', unicodedata.normalize("NFKC", str(s))).strip()

def key(name):
    k = re.sub(r'\b(dr|mr|mrs|ms|prof|professor)\.?\b', '', name.lower())
    k = re.sub(r',.*$', '', k)
    k = re.sub(r'\b(ph\.?d|ed\.?d|mba|afc|cfp|jr|sr|ii|iii)\b\.?', '', k)
    return re.sub(r'[^a-z ]', '', k).split()[-1] + '|' + (re.sub(r'[^a-z ]', '', k).split() or [''])[0]

rows = []          # dicts
seen = {}          # key -> row index

def add(**kw):
    kw.setdefault('whale', ''); kw.setdefault('score', ''); kw.setdefault('segment', '')
    kw.setdefault('address', ''); kw.setdefault('subject', ''); kw.setdefault('alt', '')
    kw.setdefault('note', ''); kw.setdefault('why', ''); kw.setdefault('role', '')
    k = key(kw['name'])
    if k in seen:                       # same person, later batch -> fold in
        r = rows[seen[k]]
        if r.get('alt'):
            r['alt'] = r['alt'] + "\n\n----- " + kw['batch'] + " version -----\n\n" + kw['msg']
        else:
            r['alt'] = kw['msg']
        bits = []
        if kw.get('subject'): bits.append('Subject: ' + kw['subject'])
        bits.append(f"Long-form version from {kw['batch']}")
        if kw.get('address'): bits.append('Email: ' + kw['address'])
        r['note'] = ((r.get('note','') + ' | ') if r.get('note') else '') + ' · '.join(bits)
        if kw['whale'] and not r['whale']: r['whale'] = kw['whale']
        if kw.get('address') and not r['address']: r['address'] = kw['address']
        return
    seen[k] = len(rows)
    rows.append(kw)

# ---------- numbered batches (### Name / *role* / *degree · channel* / ```msg```) ----------
NUMBERED = [("letters-1-198.md", "1-202"), ("letters-199-396.md", "199-396"),
            ("letters-397-516.md", "397-516"), ("letters-517-616.md", "517-616")]
for path, batch in NUMBERED:
    txt = open(path).read()
    for m in re.finditer(r'\n### ([^\n]+)\n(.*?)```\n(.*?)\n```', txt, re.S):
        head, meta, msg = m.group(1), m.group(2), m.group(3)
        whale = '🐋' if '🐋' in head else ''
        name = n(head.replace('🐋', '').replace('🔵', '').rstrip(' ,'))
        italics = re.findall(r'\*([^*\n]+)\*', meta)
        role = n(italics[0]) if italics else ''
        chan = n(italics[1]) if len(italics) > 1 else ''
        add(whale=whale, batch=batch, name=name, role=role, channel=chan,
            msg=n(msg), why=role)

# ---------- new July (## N. Name / *seg · conn · degree* / *role* / skinny / ```) ----------
txt = open("letters-new-july.md").read()
for m in re.finditer(r'\n## (\d+)\. ([^\n]+)\n(.*?)```\n(.*?)\n```', txt, re.S):
    head, meta, msg = m.group(2), m.group(3), m.group(4)
    whale = '🐋' if '🐋' in head else ''
    italics = re.findall(r'\*([^*\n]+)\*', meta)
    skinny = re.search(r'\*\*The skinny:\*\*\s*(.*)', meta)
    add(whale=whale, batch="New July", name=n(head.replace('🐋', '')),
        role=n(italics[1]) if len(italics) > 1 else '',
        channel=n(italics[0]) if italics else '',
        why=n(skinny.group(1))[:300] if skinny else '', msg=n(msg))

# ---------- new July 28-29 (same format as letters-new-july.md) ----------
txt = open("letters-new-july-28-29.md").read()
for m in re.finditer(r'\n## (\d+)\. ([^\n]+)\n(.*?)```\n(.*?)\n```', txt, re.S):
    head, meta, msg = m.group(2), m.group(3), m.group(4)
    whale = '🐋' if '🐋' in head else ''
    italics = re.findall(r'\*([^*\n]+)\*', meta)
    skinny = re.search(r'\*\*The skinny:\*\*\s*(.*)', meta)
    add(whale=whale, batch="New Jul 28-29", name=n(head.replace('🐋', '')),
        role=n(italics[1]) if len(italics) > 1 else '',
        channel=n(italics[0]) if italics else '',
        why=n(skinny.group(1))[:300] if skinny else '', msg=n(msg))

# ---------- MAESP / NAESP + Delaware (## or # N. Name / **Email:** / **Subject:** / ```) ----
def parse_email_file(path, batch, hlevel):
    txt = open(path).read()
    pat = r'\n' + ('#' * hlevel) + r' ([^\n]*?\d+\.[^\n]+)\n(.*?)```\n(.*?)\n```'
    for m in re.finditer(pat, txt, re.S):
        head, meta, msg = m.group(1), m.group(2), m.group(3)
        whale = '🔵' if '🔵' in head else ('🐋' if '🐋' in head else '')
        name = n(re.sub(r'^.*?\d+\.\s*', '', head.replace('🐋', '').replace('🔵', '')))
        name = re.sub(r'\s*—\s*PHONE SCRIPT.*$', '', name, flags=re.I)
        em = re.search(r'\*\*Email:\*\*\s*`?([^`\n*]+)`?', meta)
        sub = re.search(r'\*\*Subject:\*\*\s*`([^`]+)`', meta)
        role = re.search(r'\*\*([^*\n]+)\*\*', meta)
        when = re.search(r'\*\*When:\*\*\s*([^\n]+)', meta)
        add(whale=whale, batch=batch, name=name,
            role=n(role.group(1)) if role else '',
            address=(n(em.group(1)) if em
                     else ('302-674-0630 (no email published — call)' if 'Brown' in name else '')),
            subject=n(sub.group(1)) if sub else '',
            channel=('Phone' if 'PHONE SCRIPT' in head.upper()
                     else 'Phone / web form' if re.search(r'none published|no published|no author-application contact', meta, re.I)
                     else 'Email'), msg=n(msg),
            note=n(when.group(1)) if when else '')

parse_email_file("letters-maesp-naesp.md", "MAESP/NAESP", 2)
parse_email_file("letters-delaware.md",  "Delaware",    1)

# ---------- Mia + Rachel standalone ----------
t = open("letter-mia-russell.md").read()
blocks = re.findall(r'```\n(.*?)\n```', t, re.S)
sub = re.search(r'\*\*Subject:\*\*\s*`([^`]+)`', t)
add(whale='🐋', batch="Standalone", name="Mia Baytop Russell",
    role="Senior Lecturer, Johns Hopkins; Dir., Clark Scholars. Ex-Money Smart (UMD Ext.), ex-VP Wells Fargo Foundation",
    channel="Connection note, then InMail", subject=n(sub.group(1)) if sub else '',
    msg=n(blocks[0]), alt=n(blocks[1]) if len(blocks) > 1 else '',
    why="Practitioner, funder and academic in one person. Reference-letter candidate for Jump$tart.",
    note="Col M = 280-char connection note. Col N = 1,323-char InMail.")

t = open("letter-edoho-eket.md").read()
blocks = re.findall(r'```\n(.*?)\n```', t, re.S)
add(whale='🐋', batch="Standalone", name="Dr. Rachel Edoho-Eket",
    role="Principal, Waterloo Elementary (Blue Ribbon); Past President, MAESP; 2x author",
    channel="LinkedIn DM (1st-degree)", msg=n(blocks[0]),
    why="Maryland elementary gatekeeper, local, and an author who judges production.",
    note="Already a connection — single DM, no request note.")

# ---------- Top 25 week 1 (## N. Name — score / *role* / ```) ----------
txt = open("letters-top25-week1.md").read()
for m in re.finditer(r'\n## (\d+)\. ([^\n]+?) — (\d+)\n\*([^*\n]+)\*(.*?)```\n(.*?)\n```', txt, re.S):
    num, nm, score, role, meta, msg = m.groups()
    whale = '🐋' if '🐋' in nm else ''
    name = n(nm.replace('🐋','').replace('🔵',''))
    chan = 'InMail' if 'InMail' in role else 'LinkedIn message (1st-degree)'
    sub = re.search(r'\*\*Subject:\*\*\s*`([^`]+)`', meta)
    add(whale=whale, batch="TOP25 wk1", name=name, role=n(role), score=int(score),
        channel=chan, subject=n(sub.group(1)) if sub else '',
        why=f"Week-1 top 25, rank {num}, score {score}. Full-length rewrite using the real channel limit.",
        msg=n(msg))

# ---------- CFPB / NCLC consumer-rights (## N. Name / *chan* / ```) ----------
txt = open("letters-cfpb-consumer-rights.md").read()
for m in re.finditer(r'\n## \d+\. ([^\n]+)\n\*([^*\n]+)\*(.*?)```\n(.*?)\n```', txt, re.S):
    head, chan, meta, msg = m.group(1), m.group(2), m.group(3), m.group(4)
    whale = '🐋' if '🐋' in head else ''
    name = n(head.replace('🐋', '').replace('🔵', ''))
    add(whale=whale, batch="CFPB/rights", name=name, channel=n(chan.split('·')[0]),
        why="Kids-as-consumers-with-rights thesis. " + n(re.sub(r'\*\*[^*]+\*\*','',meta))[:220],
        msg=n(msg))

# ---------- Jump$tart / Stanford (## X. Name — ... / **Subject:** / ```) ----------
txt = open("jumpstart-clearinghouse-strategy.md").read()
for m in re.finditer(r'\n## ([A-G])\. ([^\n]+)\n(.*?)```\n(.*?)\n```', txt, re.S):
    letter, head, meta, msg = m.groups()
    whale = '🔵' if '🔵' in head else ('🐋' if '🐋' in head else '')
    name = n(re.split(r'\s+—\s+', head.replace('🐋', '').replace('🔵', ''))[0])
    if letter == 'C': name = 'Reference letter — draft for Wally to sign'
    em = re.search(r'\*\*Email:\*\*\s*`([^`]+)`', meta)
    sub = re.search(r'\*\*Subject:\*\*\s*`([^`]+)`', meta)
    purpose = n(re.sub(r'^[^—]*—\s*', '', head)) if '—' in head else ''
    add(whale=whale, batch="Jump$tart", name=name, role=purpose,
        address=n(em.group(1)) if em else '', channel='Email',
        subject=n(sub.group(1)) if sub else '', msg=n(msg), why=purpose)

# ================= workbook =================
BLUE, ORANGE, CREAM = "0054A6", "FF6B2B", "F5ECD7"
thin = Side(style='thin', color='D9D2BF')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ---- Master ----
ws = wb.active; ws.title = "Master"
HDR = ["🐋", "Batch", "Name", "Role / Org", "Why they matter", "Channel",
       "Address", "Subject", "Chars", "Message (ready to paste)",
       "Alt / long-form", "Notes", "Sent?", "Reply?"]
ws.append(HDR)
for r in rows:
    ws.append([r['whale'], r['batch'], r['name'], r.get('role', ''), r.get('why', ''),
               r.get('channel', ''), r.get('address', ''), r.get('subject', ''),
               len(r['msg']), r['msg'], r.get('alt', ''), r.get('note', ''), "", ""])
widths = [4, 11, 26, 34, 40, 20, 30, 40, 7, 78, 60, 34, 8, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORD
    if row[0].value:
        row[0].fill = PatternFill("solid", fgColor="FFE7D6" if row[0].value == '🐋' else "D6E8FF")
style_header(ws, len(HDR))

# ---- Whales ----
wsw = wb.create_sheet("Whales")
wsw.append(HDR)
for r in rows:
    if r['whale']:
        wsw.append([r['whale'], r['batch'], r['name'], r.get('role', ''), r.get('why', ''),
                    r.get('channel', ''), r.get('address', ''), r.get('subject', ''),
                    len(r['msg']), r['msg'], r.get('alt', ''), r.get('note', ''), "", ""])
for i, w in enumerate(widths, 1):
    wsw.column_dimensions[get_column_letter(i)].width = w
for row in wsw.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BORD
    row[0].fill = PatternFill("solid", fgColor="FFE7D6" if row[0].value == '🐋' else "D6E8FF")
style_header(wsw, len(HDR))

# ---- Deadlines ----
wsd = wb.create_sheet("Deadlines")
DH = ["Priority", "What", "Organization", "Hard date", "Cost", "Contact", "Why / note", "Done?"]
wsd.append(DH)
DEAD = [
 ["1","Mail one copy for Best Children's Books of the Year","Bank Street College Children's Book Committee","Nov 1, 2026","Postage","bookcom@bankstreet.edu","Deadline is Nov 1 of the publication year. No stated bar on self-published titles. Cheapest credential on any list.","",],
 ["2","Ask about associate membership before the annual meeting","Educational Book & Media Association (EBMA)","January meeting","Dues not published — ask","admin@edupaperback.org · 540-318-7770","One-on-One Sessions: every wholesaler meets every publisher member for eight minutes. No other route into that room.","",],
 ["3","Get dates + Discovery Hall exhibitor pricing","International Conference on ADHD (ACO/ADDA/CHADD)","2026 — Baltimore","Ask","301-306-7070 ext. 109","National ADHD conference in your city. CHADD already publishes on kids, impulse control and spending.","",],
 ["4","Clearinghouse eligibility question, then submit toolkit","Jump$tart Clearinghouse","Rolling","Free","suzann.knight@jumpstart.org","Submit the FREE TOOLKIT first, not the book — it answers the nationwide-availability criterion outright.","",],
 ["5","Nomination — book of the year","EIFLE Awards (Institute for Financial Literacy)","Mid-January","$75","EIFLE@FinancialLit.org","Ask Wally to nominate. He took the 2026 Corey Carlisle Award and 2026 EIFLE Educator of the Year.","",],
 ["6","Request National Partner application","Jump$tart Coalition","Quarterly review","Dues by org size","jumpstart.org/about/partners","Same reference letters as the Clearinghouse. Quarterly clock punishes waiting.","",],
 ["7","Submit — Children's Books category","NAPPA Awards","Year-round","Fee — ask Elena","elena.epstein@nappaawards.com","Self-published accepted, any release year, two physical samples, 6–8 weeks.","",],
 ["8","Pitch 'Editor's Picks: We Love It!'","Baltimore's Child","Rolling","Free","410-902-2300","Home market, standing book column, costs a phone call.","",],
 ["9","Register an event","Money Smart Week (ALA + Chicago Fed)","April","Free","moneysmartweek.org","Their own topic list already includes saving money through couponing. Pair with the Rehoboth and Lewes library letters.","",],
 ["10","Pitch the holiday gift guide","Regional parenting magazines (PMA members)","Pitch in September","Free","parentmedia.org directory","December issues are built in early fall. Miss it and it costs a full year.","",],
 ["11","Submit a 500–800 word parent-voice post","ADDitude Magazine","Rolling","Free","submissions@additudemag.com","They take posts from parents, not press releases. Book goes in the byline.","",],
 ["12","Ask whether CBC membership is required to submit","NCSS / CBC Notable Social Studies Trade Books","Submissions open spring 2027","Unknown","laura.peraza@cbcbooks.org","Best-fit list in existence. Settle the membership question now, not in 2027.","",],
 ["13","Apply to the K–12 Teaching Artist Roster","Arts for Learning Maryland","Rolling","Free","artsforlearningmd.org","The one lane that pays. Author + illustrator + finance attorney is a combination nobody else on that roster has.","",],
 ["14","Ask about the April event","Beachside Book Festival (Sussex County Libraries + Browseabout)","April — ask in winter","Free/TBD","Sussex County Libraries","Local author talks in Rehoboth. Baldacci headlined 2025, so the traffic is real.","",],
 ["15","Propose a session","NPEN monthly PD webinar series","Rolling","Free","npen.org","Twelve slots a year need filling. Audience trains the people who teach parents.","",],
 ["16","Sponsor-a-classroom-set pitch","Credit unions / National Credit Union Foundation","No deadline","—","ncuf.coop","The Berenstain Bears program proves credit unions buy children's book series for elementary schools. CRA memo already written.","",],
 ["—","DEAD — do not enter","Mathical Book Prize","—","—","—","Self-published titles are not eligible.","n/a"],
 ["—","DEAD — do not enter","Parents' Choice Awards","—","—","—","Foundation closed July 31, 2022. Still on every listicle.","n/a"],
 ["—","DEAD — window closed","Junior Library Guild","—","—","—","Wants submissions 4–6 months BEFORE publication. Keep for book two.","n/a"],
 ["—","DEAD — wrong ages","Reach Out and Read","—","—","—","6 months through age 5. Your floor is 6.","n/a"],
]
for d in DEAD: wsd.append(d)
for i, w in enumerate([9, 46, 44, 24, 20, 38, 74, 8], 1):
    wsd.column_dimensions[get_column_letter(i)].width = w
for row in wsd.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BORD
    if row[0].value == '—':
        for c in row: c.fill = PatternFill("solid", fgColor="F2F2F2"); c.font = Font(color="8A7F68")
style_header(wsd, len(DH))

# ---- Mail list ----
wsm = wb.create_sheet("Physical copies")
MH = ["Tier", "Who", "Why physical matters", "Mailed?"]
wsm.append(MH)
MAIL = [
 ["1 — mail on a yes", "Dr. Rachel Edoho-Eket", "Connected, local, Blue Ribbon principal, past MAESP president, and a two-time author. Authors judge production.", ""],
 ["1 — mail on a yes", "Dr. Christopher Wooleyhand", "Runs MAESP and a building. Can test it and distribute it.", ""],
 ["1 — mail on a yes", "Sandy Spavone", "FCCLA, 257,000 members. A chapter service project starts with her holding it.", ""],
 ["1 — mail on a yes", "Mia Baytop Russell", "Ran Money Smart at Maryland Extension, funded youth financial ed at Wells Fargo, teaches at Hopkins. Local.", ""],
 ["1 — mail on a yes", "Kaylen Tucker", "Editor of Principal. Publications cannot review a link.", ""],
 ["1 — mail on a yes", "Sarah Wright (NYPL), Amy Boglarski (Boston PL)", "Librarians assess binding, trim size and shelf-fit by hand.", ""],
 ["1 — mail on a yes", "Neale Godfrey, Cinders McLeod", "Peer authors. The offer is a trade, not a gift — ask for theirs back.", ""],
 ["DE — mail early", "Dr. Bonnie Meszaros 🔵", "The exception to mail-on-a-yes. Co-authored the National Standards for Financial Literacy; built a K–1 curriculum where every lesson wraps a picture book. Judging whether a children's book carries real economics is her job.", ""],
 ["DE", "Dr. Jennifer Nauman", "Superintendent over five elementary schools in Lewes, Rehoboth and Milton. Eight years an elementary principal.", ""],
 ["DE", "Equetta Jones", "Love Creek principal, moderator of NAESP's Early Career Principal Community of Practice.", ""],
 ["DE — hand deliver", "Matthew Keen", "New principal at Rehoboth Elementary. Four blocks from the boardwalk and nobody walks a book into a school.", ""],
 ["2 — if they engage", "Tracy Hilliard · Justin Holbrook · Megan Kirts · Daniel Rauen · Annie Shoen · Nicolle Hood · Devon Abejo · Elissa Dee · Dante Rhodes · Linda Fussell", "Second wave.", ""],
 ["Do not mail", "Podcasters, reporters, columnists, academics", "They want a hook, a fact, or a searchable PDF. A parcel is friction. Includes the Stanford three — Lusardi, Sticha, Mahoney: digital only.", "n/a"],
]
for m_ in MAIL: wsm.append(m_)
for i, w in enumerate([20, 46, 92, 9], 1):
    wsm.column_dimensions[get_column_letter(i)].width = w
for row in wsm.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BORD
style_header(wsm, len(MH))

# ---- 86ed ----
ws8 = wb.create_sheet("86ed")
ws8.append(["Name", "Disposition", "Reason"])
try:
    old = openpyxl.load_workbook("CGB_NEW_1st_degree_July.xlsx", read_only=True)
    src = old["86ed + already covered"]
    for r in src.iter_rows(min_row=2, values_only=True):
        if r and r[0]: ws8.append(list(r))
    old.close()
except Exception as e:
    print("86ed carry-over skipped:", e)
for i, w in enumerate([30, 18, 96], 1):
    ws8.column_dimensions[get_column_letter(i)].width = w
for row in ws8.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = BORD
style_header(ws8, 3)

# ---- README sheet ----
wsr = wb.create_sheet("Read me", 0)
wsr.column_dimensions['A'].width = 118
NOTES = [
 ("Clarence Gets a Bargain — master outreach workbook", True),
 ("", False),
 ("Rebuilt from the markdown sources. This supersedes the five earlier CGB_HVT workbooks.", False),
 ("", False),
 ("Master — every letter, one row per person. Column J is the paste-ready message.", False),
 ("Whales — the same rows, filtered. 🐋 = whale. 🔵 = blue whale, mail a copy.", False),
 ("Deadlines — the dated actions. Sort by column A. Grey rows are dead ends, kept so nobody re-finds them.", False),
 ("Physical copies — who gets a book, and the mail-on-a-yes rule.", False),
 ("86ed — dispositions carried over.", False),
 ("", False),
 ("Where one person appears in more than one batch there is ONE row. The short LinkedIn note", False),
 ("is in column J; the long-form email is in column K; column L says which file it came from.", False),
 ("", False),
 ("Sent? and Reply? are yours. Nothing in this rebuild overwrites them — they were empty.", False),
 ("", False),
 ("Copying out of cells is still worse than copying out of the markdown. The .md files remain", False),
 ("the working copy; this workbook is the tracker.", False),
]
for i, (t, bold) in enumerate(NOTES, 1):
    c = wsr.cell(row=i, column=1, value=t)
    c.font = Font(bold=bold, size=13 if bold else 10, color=BLUE if bold else "111111")
    c.alignment = Alignment(vertical="top", wrap_text=True)

wb.save("CGB_MASTER_outreach.xlsx")

# report
from collections import Counter
print("rows:", len(rows), "| whales:", sum(1 for r in rows if r['whale']))
print(Counter(r['batch'] for r in rows))
print("folded duplicates:", sum(1 for r in rows if r.get('alt')))
for r in rows:
    if r.get('alt'): print("   ", r['name'], "->", r.get('note', '')[:70])
