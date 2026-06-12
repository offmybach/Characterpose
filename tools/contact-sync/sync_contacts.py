#!/usr/bin/env python3
"""Append new LinkedIn connections to an existing contacts workbook.

Reads rows from a CSV (a LinkedIn "Connections.csv" export, or any CSV with
recognizable headers) or from JSON, dedupes them against the workbook, fills in
whatever columns the workbook already has, tags each contact with a CGB
audience group, and appends the new rows.

The workbook is the source of truth. A timestamped backup is written before
every save, so a bad run can never destroy the categorized data. Use --dry-run
to preview exactly what would change before writing anything.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date
from pathlib import Path
import shutil

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Run:  pip install openpyxl")

# Default target on Jonathan's machine; override with --file anywhere else.
DEFAULT_FILE = r"C:\Users\mdmen\Downloads\finlit_contacts_categorized_bespoke_groups.xlsx"

# Map a spreadsheet header to a canonical field. The first canonical whose any
# substring appears in the (lowercased) header wins, so order is deliberate:
# the specific columns sit ahead of the generic "name" catch so that, e.g.,
# "Company Name" lands on company, not full_name.
HEADER_PATTERNS = [
    ("first_name",   ["first name", "first"]),
    ("last_name",    ["last name", "last"]),
    ("email",        ["email", "e-mail"]),
    ("profile_url",  ["url", "linkedin", "profile"]),
    ("company",      ["company", "organization", "organisation", "employer", "org"]),
    ("title",        ["position", "title", "role", "headline", "job"]),
    ("location",     ["location", "city", "region", "area"]),
    ("connected_on", ["connected", "connection date"]),
    ("added_on",     ["added", "imported", "synced"]),
    ("group",        ["group", "category", "segment", "bucket", "audience", "bespoke", "tag", "list"]),
    ("notes",        ["note", "why", "pitch", "comment", "opener", "message"]),
    ("full_name",    ["full name", "name"]),
]

# Ordered classification rules. First rule whose any keyword appears in the
# combined title+company+headline text wins. Front-line school staff beat
# org-type words so a "teacher at X School District" reads as Educator.
SEGMENT_RULES = [
    ("Librarian", [
        "librarian", "library", "media specialist", "media center", "school library",
    ]),
    ("Educator (K-5)", [
        "teacher", "elementary", "kindergarten", "classroom", "paraprofessional",
        "reading specialist", "literacy coach", "instructional coach", "k-5", "k–5",
        "special education teacher", "ela teacher", "primary school", "grade teacher",
    ]),
    ("Institutional / Bulk", [
        "credit union", "bank", "financial literacy", "financial education", "fdic",
        "ncua", "foundation", "nonprofit", "non-profit", "school district", "district",
        "superintendent", "curriculum director", "county office", "treasurer",
        "chamber of commerce", "united way", "extension office", "university",
        "college", "government", "municipal", "529", "principal", "administrator",
    ]),
    ("Educator (K-5)", [
        "educator", "education", "teaching", "school", "curriculum", "instruction",
        "stem", "ela", "tutor",
    ]),
]
DEFAULT_SEGMENT = "Parent / General"

# Roots used to reuse the sheet's own existing group label instead of imposing
# new vocabulary (so "Educators" in the sheet stays "Educators").
SEGMENT_ROOTS = {
    "Librarian": ["librar"],
    "Educator (K-5)": ["educat", "teach", "school", "k-5", "k–5"],
    "Institutional / Bulk": ["institu", "bulk", "district", "credit", "bank", "financial", "org", "bulk"],
    "Parent / General": ["parent", "family", "general", "consumer"],
}

CSV_HEADER_HINTS = [
    "first name", "last name", "url", "email address", "company", "position",
    "connected on", "headline", "profile", "name",
]


def build_header_map(headers):
    """Return [(col_index_0based, canonical_or_None), ...] for a header row."""
    mapping = []
    for i, h in enumerate(headers):
        key = str(h or "").strip().lower()
        canon = None
        if key:
            for field, subs in HEADER_PATTERNS:
                if any(s in key for s in subs):
                    canon = field
                    break
        mapping.append((i, canon))
    return mapping


def first_col(header_map, field):
    for idx, canon in header_map:
        if canon == field:
            return idx
    return None


def normalize_record(d):
    """Coerce an arbitrary input dict into canonical fields."""
    g = lambda *keys: next((str(d[k]).strip() for k in keys if d.get(k) not in (None, "")), "")
    first = g("first_name", "first")
    last = g("last_name", "last")
    full = g("full_name", "name")
    if full and not (first or last):
        parts = full.split()
        first = parts[0]
        last = " ".join(parts[1:])
    if not full and (first or last):
        full = f"{first} {last}".strip()
    return {
        "full_name": full,
        "first_name": first,
        "last_name": last,
        "email": g("email"),
        "profile_url": g("profile_url", "url", "linkedin"),
        "company": g("company", "organization", "employer"),
        "title": g("title", "position", "role", "headline", "job"),
        "location": g("location", "city", "region"),
        "connected_on": g("connected_on", "connected"),
    }


def record_key(rec):
    """A stable dedupe key: profile URL if we have one, else name+company."""
    url = rec.get("profile_url", "").strip().lower()
    if url:
        url = url.split("?")[0].rstrip("/")
        return f"url:{url}"
    name = rec.get("full_name", "").strip().lower()
    company = rec.get("company", "").strip().lower()
    if name:
        return f"nc:{name}|{company}"
    return ""


def classify_segment(rec):
    text = " ".join([rec.get("title", ""), rec.get("company", ""), rec.get("full_name", "")]).lower()
    for label, keywords in SEGMENT_RULES:
        if any(k in text for k in keywords):
            return label
    return DEFAULT_SEGMENT


def map_group_value(label, existing_values):
    """Reuse the sheet's own label for this segment when one exists."""
    roots = SEGMENT_ROOTS.get(label, [])
    for ev in existing_values:
        low = ev.lower()
        if any(r in low for r in roots):
            return ev
    return label


def read_csv_records(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return []
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        joined = " | ".join(str(c).lower() for c in row)
        if any(h in joined for h in CSV_HEADER_HINTS):
            header_idx = i
            break
    headers = rows[header_idx]
    hmap = build_header_map(headers)
    records = []
    for row in rows[header_idx + 1:]:
        if not any(str(c).strip() for c in row):
            continue
        raw = {}
        for idx, canon in hmap:
            if canon and idx < len(row):
                raw[canon] = row[idx]
        records.append(normalize_record(raw))
    return records


def read_json_records(path):
    text = sys.stdin.read() if path == "-" else Path(path).read_text(encoding="utf-8")
    data = json.loads(text)
    if isinstance(data, dict):
        data = data.get("contacts", data.get("rows", [data]))
    return [normalize_record(d) for d in data]


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", default=DEFAULT_FILE, help="path to the .xlsx workbook")
    ap.add_argument("--sheet", default=None, help="sheet/tab name (default: active sheet)")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--from-csv", metavar="PATH", help="LinkedIn export or any CSV of new contacts")
    src.add_argument("--from-json", metavar="PATH", help="JSON list of contacts ('-' for stdin)")
    ap.add_argument("--limit", type=int, default=50, help="max new rows to add in one run")
    ap.add_argument("--dry-run", action="store_true", help="preview only; write nothing")
    ap.add_argument("--no-backup", action="store_true", help="skip the pre-write backup (not recommended)")
    args = ap.parse_args()

    target = Path(args.file)
    if not target.exists():
        sys.exit(f"Workbook not found: {target}\nPass the right path with --file.")

    records = read_csv_records(args.from_csv) if args.from_csv else read_json_records(args.from_json)
    if not records:
        print("No input contacts found. Nothing to do.")
        return

    wb = load_workbook(target)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = [c.value for c in ws[1]]
    hmap = build_header_map(headers)

    mapped = [(h, canon) for (i, canon), h in zip(hmap, headers) if canon]
    group_idx = first_col(hmap, "group")
    existing_group_values = []
    if group_idx is not None:
        seen_vals = set()
        for row in ws.iter_rows(min_row=2):
            v = row[group_idx].value
            if v and str(v).strip() and str(v).strip() not in seen_vals:
                seen_vals.add(str(v).strip())
                existing_group_values.append(str(v).strip())

    # Existing keys from the sheet plus a state file, so re-adding is impossible
    # even if a row was later deleted.
    state_path = target.with_name(".contact-sync-state.json")
    seen_keys = set()
    if state_path.exists():
        try:
            seen_keys.update(json.loads(state_path.read_text(encoding="utf-8")).get("keys", []))
        except (json.JSONDecodeError, OSError):
            pass
    for row in ws.iter_rows(min_row=2):
        raw = {}
        for idx, canon in hmap:
            if canon:
                raw[canon] = row[idx].value
        k = record_key(normalize_record({kk: ("" if vv is None else vv) for kk, vv in raw.items()}))
        if k:
            seen_keys.add(k)

    new_records, skipped, batch_keys = [], 0, set()
    for rec in records:
        k = record_key(rec)
        if not k or k in seen_keys or k in batch_keys:
            skipped += 1
            continue
        batch_keys.add(k)
        new_records.append(rec)
        if len(new_records) >= args.limit:
            break

    print(f"Workbook : {target}")
    print(f"Sheet    : {ws.title}  ({ws.max_row - 1} existing rows)")
    print("Columns  : " + ", ".join(f"{h}→{c}" for h, c in mapped) or "(none recognized)")
    if group_idx is None:
        print("Note     : no Group/Category column found — contacts will be added without a CGB tag.")
    elif existing_group_values:
        print("Groups   : reusing existing labels → " + ", ".join(existing_group_values))
    print(f"Input    : {len(records)} contacts  |  new: {len(new_records)}  |  duplicates skipped: {skipped}")

    tally = {}
    today = date.today().isoformat()
    rows_to_write = []
    for rec in new_records:
        seg = classify_segment(rec)
        tally[seg] = tally.get(seg, 0) + 1
        row_values = list(headers)  # length template
        out = [None] * len(headers)
        for idx, canon in hmap:
            if canon == "group" and group_idx is not None:
                out[idx] = map_group_value(seg, existing_group_values)
            elif canon == "added_on":
                out[idx] = today
            elif canon and canon != "notes":
                val = rec.get(canon, "")
                out[idx] = val if val else None
        rows_to_write.append((rec, out))

    if tally:
        print("Tagging  : " + ", ".join(f"{k}: {v}" for k, v in tally.items()))
    for rec, _ in rows_to_write:
        who = rec["full_name"] or rec["profile_url"] or "(unnamed)"
        extra = f" — {rec['title']}" if rec["title"] else ""
        print(f"   + {who}{extra}")

    if not rows_to_write:
        print("\nNothing new to add.")
        return

    if args.dry_run:
        print(f"\nDRY RUN: would add {len(rows_to_write)} row(s). No file written.")
        return

    if not args.no_backup:
        backups = target.parent / "contact-sync-backups"
        backups.mkdir(exist_ok=True)
        stamp = date.today().isoformat()
        backup = backups / f"{target.stem}_{stamp}.xlsx"
        n = 1
        while backup.exists():
            backup = backups / f"{target.stem}_{stamp}_{n}.xlsx"
            n += 1
        shutil.copy2(target, backup)
        print(f"\nBackup   : {backup}")

    for _, out in rows_to_write:
        ws.append(out)
    try:
        wb.save(target)
    except PermissionError:
        sys.exit(f"\nCould not save {target}.\nClose the file in Excel and run again.")

    for rec in new_records:
        k = record_key(rec)
        if k:
            seen_keys.add(k)
    state_path.write_text(json.dumps({"keys": sorted(seen_keys)}, indent=0), encoding="utf-8")
    print(f"Added    : {len(rows_to_write)} row(s). Done.")


if __name__ == "__main__":
    main()
