#!/usr/bin/env python3
"""Append new LinkedIn connections to the FinLit contacts workbook.

This is tuned to the real workbook structure:

  * Master            — the deduped source of truth (one row per contact)
  * category tabs     — Educators Schools, Media Journalists, Financial Literacy,
                        Government Policy, Nonprofit Community, Financial Services,
                        PTA Parent Leaders, Authors Books, Other Review
  * Bespoke Email Groups — the taxonomy legend (group -> shared angle + slot)

For each new contact it classifies a Primary Category (-> Outreach Segment, tab,
and Fit Note), assigns a Bespoke Email Group, copies that group's Shared Angle and
Personalization Slot from the legend, parses credentials off the name, then
appends the row to Master and to the matching category tab.

It fills the *categorization* columns. The judgment columns — Custom Email
Priority/Score and the per-person Suggested Custom Angle / High Value Note — are
left blank for you to review (that's the tabled bespoke-email step). New rows are
marked in Parse Notes as auto-added so they are easy to find and check.

Safety: the workbook is backed up before every write. Use --dry-run to preview.
Note: openpyxl may not preserve charts/images on save, so keep the backups and
eyeball the file after the first real run.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Run:  pip install openpyxl")

DEFAULT_FILE = r"C:\Users\mdmen\Downloads\finlit_contacts_categorized_bespoke_groups.xlsx"
MASTER_SHEET = "Master"
LEGEND_SHEET = "Bespoke Email Groups"

# Primary Category -> (Outreach Segment, tab sheet name, Fit Note). Exact strings
# pulled from the workbook so new rows match the existing vocabulary.
CATEGORY_TABLE = {
    "Financial Literacy & Financial Education": ("Financial Literacy", "Financial Literacy", "direct financial literacy, money education, or consumer education fit"),
    "Educators & Schools":                      ("Educator", "Educators Schools", "school, educator, curriculum, or academic fit"),
    "Media & Journalists":                      ("Media", "Media Journalists", "media, journalist, editor, producer, or finance-news fit"),
    "Government, Policy & Regulators":          ("Government / Policy", "Government Policy", "public-sector, policy, or regulator fit"),
    "Financial Services, Fintech & Corporate":  ("Financial Services / Corporate", "Financial Services", "financial services, fintech, credit, banking, or corporate finance fit"),
    "Nonprofit & Community Programs":           ("Nonprofit / Community", "Nonprofit Community", "nonprofit, philanthropy, community, or impact fit"),
    "PTA & Parent Leaders":                     ("PTA / Parent Leader", "PTA Parent Leaders", "parent, family, or PTA leadership fit"),
    "Authors, Books & Publishing":              ("Books / Publishing", "Authors Books", "books, authorship, or publishing fit"),
    "Other / Review":                           ("Review", "Other Review", "not enough role/account signal to classify confidently"),
}

# Ordered classifier. First category whose any keyword appears in title+account
# wins. Financial Literacy sits ahead of Educators so "Financial Educator" reads
# as finlit; Nonprofit sits ahead of Financial Services so a foundation lead at a
# credit union reads as nonprofit — both matching how the sheet was built.
CLASSIFY_RULES = [
    ("Authors, Books & Publishing", ["author", "co-author", "publisher", "publishing", "novelist", "illustrator", "storytell", "children's book", "childrens book", "picture book"]),
    ("Media & Journalists", ["journalist", "editor", "reporter", "producer", "podcast", "anchor", "correspondent", "columnist", "broadcast", "newsroom", "staff writer", "host of", "media", "cnbc", "npr", "bloomberg", "forbes", "yahoo finance", "marketwatch"]),
    ("Government, Policy & Regulators", ["federal reserve", "reserve bank", "cfpb", "fdic", "occ ", "ncua", "treasury", "regulator", "policy", "public sector", "municipal", "county of", "state of", "senator", "congress", "legislat", "commissioner", "department of", "government"]),
    ("PTA & Parent Leaders", ["pta", "pto ", "parent leader", "parent-teacher", "family advocate", "parent advocate", "room parent", "parenting"]),
    ("Nonprofit & Community Programs", ["nonprofit", "non-profit", "foundation", "united way", "philanthrop", "community program", "community impact", "ngo", "charitable", "executive director"]),
    ("Financial Literacy & Financial Education", ["financial literacy", "financial education", "financial educator", "financial wellness", "financial coach", "money coach", "financial counselor", "accredited financial counselor", "afc®", "cfei", "financial capability", "financial empowerment", "personal finance educat"]),
    ("Educators & Schools", ["teacher", "educator", "professor", "instructor", "principal", "superintendent", "curriculum", "classroom", "elementary", "middle school", "high school", "k-12", "k12", "school district", "dean", "faculty", "lecturer", "paraprofessional", "school", "university", "college", "academic"]),
    ("Financial Services, Fintech & Corporate", ["financial advisor", "wealth", "financial planner", "cfp", "banker", "bank", "credit union", "fintech", "insurance", "investment", "portfolio", "broker", "financial services", "loan officer", "mortgage", "corporate"]),
]
DEFAULT_CATEGORY = "Other / Review"

# Default Bespoke Email Group per category (a best guess the user can adjust).
# Financial Services and Other / Review have no clean group, so they're left for
# review rather than forced into the wrong template.
def derive_group(category, text):
    if category == "Educators & Schools":
        return "Classroom, school, curriculum, and library practitioners"
    if category == "Media & Journalists":
        return "Personal-finance media and broad-audience storytellers"
    if category == "Authors, Books & Publishing":
        return "Children's books, publishing, and educational storytelling"
    if category == "PTA & Parent Leaders":
        return "Parent, PTA, and family advocates"
    if category == "Nonprofit & Community Programs":
        return "Community impact and nonprofit program leaders"
    if category == "Government, Policy & Regulators":
        if "federal reserve" in text or "reserve bank" in text:
            return "Federal Reserve and policy/institutional connectors"
        return "Public-sector financial education / consumer protection"
    if category == "Financial Literacy & Financial Education":
        org_words = ["coalition", "council", "curriculum", "organization", "program",
                     "initiative", "institute", "endowment", "network", "association", "foundation"]
        if any(w in text for w in org_words):
            return "Financial-literacy coalitions and curriculum organizations"
        return "Financial wellness educators, coaches, and practitioners"
    return ""  # Financial Services / Review -> leave for manual grouping

CREDENTIAL_HINTS = {
    "phd", "ph.d", "ph.d.", "mba", "ma", "ms", "msc", "med", "m.ed", "m.ed.", "cfp",
    "cfei", "afc", "clu", "cpa", "edd", "ed.d", "jd", "md", "cfa", "cfle", "apfi",
    "cflp", "ipa", "cpffe", "cude", "aif", "chfc", "ricp", "cepf", "cfp®",
}


def split_name(raw):
    """Return (display, primary_name, credentials) from a messy display name."""
    display = re.sub(r"^[\s★☆✦✱*•·–—\-]+", "", str(raw or "")).strip()
    if not display:
        return "", "", ""
    if "," in display:
        primary, creds = display.split(",", 1)
        return display, primary.strip(), creds.strip()
    words = display.split()
    creds_tail = []
    while words:
        w = words[-1]
        bare = w.lower().strip(".®™")
        if "®" in w or "™" in w or bare in CREDENTIAL_HINTS:
            creds_tail.insert(0, words.pop())
        else:
            break
    primary = " ".join(words).strip() or display
    return display, primary, " ".join(creds_tail).strip()


def norm_key(name):
    s = re.sub(r"[®™★☆✦✱*]", "", str(name or "")).lower()
    s = re.sub(r"[.,]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def classify(title, account):
    text = f"{title} {account}".lower()
    for category, keywords in CLASSIFY_RULES:
        if any(k in text for k in keywords):
            return category
    return DEFAULT_CATEGORY


def fit_score(category, degree):
    if category == "Other / Review":
        return 2
    score = 3
    if str(degree).strip().lower().startswith("1st"):
        score += 1
    if category in ("Financial Literacy & Financial Education", "Educators & Schools",
                    "Media & Journalists", "PTA & Parent Leaders"):
        score += 1
    return max(1, min(5, score))


def normalize_record(d):
    g = lambda *keys: next((str(d[k]).strip() for k in keys if d.get(k) not in (None, "")), "")
    return {
        "name": g("name", "full_name", "display_name", "linkedin_display_name"),
        "title": g("title", "headline", "position", "role"),
        "account": g("account", "company", "organization", "employer"),
        "geography": g("geography", "location", "city", "region"),
        "degree": g("degree", "li_connection_degree", "connection_degree") or "1st",
        "connected_on": g("connected_on", "connected", "saved_recent_date", "date"),
    }


CSV_HEADER_HINTS = ["first name", "last name", "url", "email address", "company", "position", "connected on", "name", "headline", "title"]

CSV_FIELD_MAP = [
    ("first_name", ["first name"]), ("last_name", ["last name"]),
    ("name", ["name", "display name"]), ("title", ["position", "title", "headline"]),
    ("account", ["company", "account", "organization"]), ("geography", ["location", "geography"]),
    ("degree", ["connection", "degree"]), ("connected_on", ["connected on", "connected", "date"]),
]


def map_csv_row(headers, row):
    out = {}
    low = [str(h or "").strip().lower() for h in headers]
    for field, subs in CSV_FIELD_MAP:
        for i, h in enumerate(low):
            if any(s in h for s in subs) and i < len(row) and row[i] not in (None, ""):
                out.setdefault(field, str(row[i]).strip())
    if not out.get("name") and (out.get("first_name") or out.get("last_name")):
        out["name"] = f"{out.get('first_name','')} {out.get('last_name','')}".strip()
    return out


def read_csv_records(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return []
    hidx = 0
    for i, row in enumerate(rows[:10]):
        joined = " | ".join(str(c).lower() for c in row)
        if any(h in joined for h in CSV_HEADER_HINTS):
            hidx = i
            break
    headers = rows[hidx]
    return [normalize_record(map_csv_row(headers, r)) for r in rows[hidx + 1:] if any(str(c).strip() for c in r)]


def read_json_records(path):
    text = sys.stdin.read() if path == "-" else Path(path).read_text(encoding="utf-8")
    data = json.loads(text)
    if isinstance(data, dict):
        data = data.get("contacts", data.get("rows", [data]))
    return [normalize_record(d) for d in data]


def load_legend(wb):
    legend = {}
    if LEGEND_SHEET not in wb.sheetnames:
        return legend
    ws = wb[LEGEND_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if r and r[0]:
            legend[str(r[0]).strip()] = (str(r[5]).strip() if len(r) > 5 and r[5] else "",
                                         str(r[6]).strip() if len(r) > 6 and r[6] else "")
    return legend


def existing_keys(ws):
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    name_idx = None
    for cand in ("LinkedIn Display Name", "Name", "Primary Name / Org"):
        if cand in headers:
            name_idx = headers.index(cand)
            break
    keys = set()
    if name_idx is None:
        return keys
    for row in ws.iter_rows(min_row=2, values_only=True):
        if name_idx < len(row) and row[name_idx]:
            _, primary, _ = split_name(row[name_idx])
            k = norm_key(primary)
            if k:
                keys.add(k)
    return keys


def insert_records_top(ws, fields_list):
    """Insert new rows directly under the header (row 2), so the newest contacts
    sit at the top. Existing rows shift down; input order is preserved top-down."""
    if not fields_list:
        return
    ws.insert_rows(2, amount=len(fields_list))
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    for i, fields in enumerate(fields_list):
        for col, h in enumerate(headers, start=1):
            ws.cell(row=2 + i, column=col, value=fields.get(h))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", default=DEFAULT_FILE, help="path to the contacts .xlsx")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--from-csv", metavar="PATH", help="LinkedIn export or any CSV")
    src.add_argument("--from-json", metavar="PATH", help="JSON list of contacts ('-' for stdin)")
    ap.add_argument("--limit", type=int, default=100, help="max new rows per run")
    ap.add_argument("--no-tabs", action="store_true", help="write to Master only, skip category tabs")
    ap.add_argument("--dry-run", action="store_true", help="preview only; write nothing")
    ap.add_argument("--no-backup", action="store_true", help="skip the pre-write backup (not recommended)")
    args = ap.parse_args()

    target = Path(args.file)
    if not target.exists():
        sys.exit(f"Workbook not found: {target}\nPass the right path with --file.")

    records = read_csv_records(args.from_csv) if args.from_csv else read_json_records(args.from_json)
    if not records:
        print("No input contacts found. Nothing to do.")
        return

    wb = load_workbook(target)
    if MASTER_SHEET not in wb.sheetnames:
        sys.exit(f"No '{MASTER_SHEET}' sheet in {target.name}.")
    master = wb[MASTER_SHEET]
    legend = load_legend(wb)

    state_path = target.with_name(".contact-sync-state.json")
    seen = set()
    if state_path.exists():
        try:
            seen.update(json.loads(state_path.read_text(encoding="utf-8")).get("keys", []))
        except (json.JSONDecodeError, OSError):
            pass
    seen |= existing_keys(master)

    today = date.today().isoformat()
    planned, skipped, batch = [], 0, set()
    for rec in records:
        _, primary, creds = split_name(rec["name"])
        key = norm_key(primary)
        if not key or key in seen or key in batch:
            skipped += 1
            continue
        batch.add(key)
        category = classify(rec["title"], rec["account"])
        segment, tab, fit_note = CATEGORY_TABLE[category]
        group = derive_group(category, f"{rec['title']} {rec['account']}".lower())
        shared, slot = legend.get(group, ("", ""))
        fields = {
            "Custom Email Priority": None, "Custom Email Score": None,
            "Bespoke Email Group": group or None,
            "High Value Note": None, "Suggested Custom Angle": None,
            "Shared Bespoke Email Angle": shared or None, "Personalization Slot": slot or None,
            "Primary Category": category, "Outreach Segment": segment, "Fit Score": fit_score(category, rec["degree"]),
            "Name": rec["name"], "LinkedIn Display Name": rec["name"],
            "Primary Name / Org": primary, "Credentials / Descriptor / Org Name": creds or None,
            "Org/Profile Flag": None, "LinkedIn Source Page": None,
            "LI Connection Degree": rec["degree"], "Sales Navigator Lists": None,
            "Title": rec["title"] or None, "Account": rec["account"] or None,
            "Source Lists": "Contact sync", "Geography": rec["geography"] or None,
            "Outreach Activity": None, "Saved/Recent Date": rec["connected_on"] or today,
            "Fit Notes": fit_note, "Secondary Category Matches": None,
            "Source Screenshots": None, "Parse Notes": f"auto-added by contact-sync {today}",
            "Source Files": "contact-sync", "Duplicate Source Rows": None,
        }
        planned.append((category, tab, fields))
        if len(planned) >= args.limit:
            break

    print(f"Workbook : {target}")
    print(f"Master   : {master.max_row - 1} existing rows")
    print(f"Input    : {len(records)}  |  new: {len(planned)}  |  duplicates skipped: {skipped}")
    tally = {}
    for category, tab, fields in planned:
        tally[category] = tally.get(category, 0) + 1
        dest = "Master" if args.no_tabs else f"Master + {tab}"
        print(f"   + {fields['Primary Name / Org']}  →  {category}  [{dest}]"
              + (f"  group: {fields['Bespoke Email Group']}" if fields["Bespoke Email Group"] else "  group: (review)"))
    if tally:
        print("By category: " + ", ".join(f"{k}: {v}" for k, v in tally.items()))

    if not planned:
        print("\nNothing new to add.")
        return
    if args.dry_run:
        print(f"\nDRY RUN: would add {len(planned)} row(s) to Master"
              + ("" if args.no_tabs else " and their category tabs") + ". No file written.")
        return

    if not args.no_backup:
        backups = target.parent / "contact-sync-backups"
        backups.mkdir(exist_ok=True)
        backup = backups / f"{target.stem}_{today}.xlsx"
        n = 1
        while backup.exists():
            backup = backups / f"{target.stem}_{today}_{n}.xlsx"
            n += 1
        shutil.copy2(target, backup)
        print(f"\nBackup   : {backup}")

    insert_records_top(master, [fields for _, _, fields in planned])
    if not args.no_tabs:
        by_tab = {}
        for _, tab, fields in planned:
            if tab in wb.sheetnames:
                by_tab.setdefault(tab, []).append(fields)
        for tab, rows in by_tab.items():
            insert_records_top(wb[tab], rows)
    try:
        wb.save(target)
    except PermissionError:
        sys.exit(f"\nCould not save {target}. Close it in Excel and run again.")

    for _, _, fields in planned:
        seen.add(norm_key(fields["Primary Name / Org"]))
    state_path.write_text(json.dumps({"keys": sorted(seen)}, indent=0), encoding="utf-8")
    print(f"Added    : {len(planned)} row(s). Done.")


if __name__ == "__main__":
    main()
